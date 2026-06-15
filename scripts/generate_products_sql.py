"""Reads equipment_data_template.xlsx and emits supabase/02_products_seed.sql.

Each product is categorised from its Series (column D). Only that family's spec
columns are pulled into the JSONB `specs`, so every record carries exactly the
attributes that apply to it.

  standard (all)      A,B,C,D,E,AB
  compressor (F–Z)    Oil free: AB, OF      Oil lube: EG, EN, EQ
  air tank (AE–AI)    AT
  air filter (AK–AQ)  AF
  dryer (AS–AZ)       Dryer

Run:  python scripts/generate_products_sql.py
"""
import json
import os
import openpyxl
from openpyxl.utils import column_index_from_string as cidx, get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "equipment_data_template.xlsx")
OUT = os.path.join(ROOT, "supabase", "02_products_seed.sql")

# Series (column D) -> category
SERIES_CATEGORY = {
    "AB": "Oil free compressor", "OF": "Oil free compressor",
    "EG": "Oil lube compressor", "EN": "Oil lube compressor", "EQ": "Oil lube compressor",
    "AT": "Air Tank", "AF": "Air filter", "Dryer": "Dryer",
}
# category -> inclusive spec-column range (Excel letters)
CATEGORY_RANGE = {
    "Oil free compressor": ("F", "Z"), "Oil lube compressor": ("F", "Z"),
    "Air Tank": ("AE", "AI"), "Air filter": ("AK", "AQ"), "Dryer": ("AS", "AZ"),
}
# core/filterable columns promoted to real DB columns: db_col -> Excel letter
CORE = {
    "model": "A", "tpl": "B", "type": "C", "series": "D", "air_quality": "E",
    "price_rm": "AB", "wc_ac": "F", "kw": "G", "hp": "H", "cfm_min": "M", "cfm_max": "N",
}
# promoted letters that sit inside a spec range — don't duplicate them into specs
PROMOTED = {"F", "G", "H", "M", "N"}


def sql_str(v):
    return "null" if v is None else "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    if v in (None, ""):
        return "null"
    try:
        return str(round(float(v), 2))
    except (TypeError, ValueError):
        return "null"


def sql_money(v):
    """Prices stored as whole Ringgit (source decimals were conversion artifacts)."""
    if v in (None, ""):
        return "null"
    try:
        return str(round(float(v)))
    except (TypeError, ValueError):
        return "null"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Product"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])

    def cell(row, letter):
        i = cidx(letter) - 1
        return row[i] if i < len(row) else None

    out = ["-- Auto-generated from equipment_data_template.xlsx. Do not edit by hand.",
           "-- Re-run scripts/generate_products_sql.py to regenerate.",
           "",
           "-- Make sure the extra columns exist, then replace the catalog.",
           "alter table public.products add column if not exists category text;",
           "alter table public.products add column if not exists lead_time_weeks numeric;",
           "delete from public.products;",
           ""]

    for r in rows[1:]:
        model = cell(r, "A")
        if not model:
            continue
        series = cell(r, "D")
        category = SERIES_CATEGORY.get(series, series)  # fallback: series as-is
        rng = CATEGORY_RANGE.get(category)

        specs = {}
        if rng:
            start, end = cidx(rng[0]), cidx(rng[1])
            for ci in range(start, end + 1):
                letter = get_column_letter(ci)
                if letter in PROMOTED:
                    continue
                header = headers[ci - 1] if ci - 1 < len(headers) else None
                val = cell(r, letter)
                if header and val not in (None, ""):
                    specs[header] = val

        cols = ["model", "tpl", "type", "series", "category", "air_quality", "wc_ac",
                "kw", "hp", "cfm_min", "cfm_max", "price_rm", "specs"]
        vals = [
            sql_str(cell(r, "A")), sql_str(cell(r, "B")), sql_str(cell(r, "C")),
            sql_str(cell(r, "D")), sql_str(category), sql_str(cell(r, "E")), sql_str(cell(r, "F")),
            sql_num(cell(r, "G")), sql_num(cell(r, "H")),
            sql_num(cell(r, "M")), sql_num(cell(r, "N")),
            sql_money(cell(r, "AB")),
            sql_str(json.dumps(specs, default=str)) + "::jsonb",
        ]
        out.append(
            f"insert into public.products ({', '.join(cols)})\nvalues ({', '.join(vals)});"
        )

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(out) + "\n")
    print(f"Wrote {OUT} ({len(rows) - 1} product rows)")


if __name__ == "__main__":
    main()
